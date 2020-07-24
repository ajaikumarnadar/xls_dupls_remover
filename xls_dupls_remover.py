import openpyxl
import sys

'''tmp_arr = []
row_del = []

wb = openpyxl.load_workbook(sys.argv[1]);

#print(wb.sheetnames);

for sheet in wb:
	print(sheet);

	for i in sheet:
		cell = list(i)[2];
		val = cell.value
		if str(val) in tmp_arr:
			row_num = cell.row
			row_del.append(row_num);
		else:
			tmp_arr.append(val);
	#print(row_del[0])
	sheet.delete_rows(row_del[0])
	row_del.pop(0);

	for i in row_del:
		i=i-1
		sheet.delete_rows(i);

wb.save('new.xlsx')'''


wbr = openpyxl.load_workbook(sys.argv[1]);

wsr = wbr.worksheets[0]
wbw = openpyxl.Workbook();
wsw= wbw.worksheets[0]

mr = wsr.max_row
mc = wsr.max_column

for i in range(1, mr+1):
	for j in range(1, mc+1):
		c = wsr.cell(row = i, column = j)

		wsw.cell(row = i, column = j).value = c.value

'''for sheetr in wbr:
	print(sheetr)
	for cellr in sheetr.iter_rows():
		print(list(cellr))
		for sheetw in wbw:
			for cellw in sheetw.iter_rows():
				#print(cellr);
				print(cellr.value);
				cellw.value=cellr.value'''



wbw.save('new.xlsx')